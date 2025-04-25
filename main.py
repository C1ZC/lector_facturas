import sys
import os
import re
import pdfplumber
from openpyxl import Workbook, load_workbook
from PyQt5.QtWidgets import (QApplication, QMainWindow, QPushButton, QFileDialog, 
                           QLabel, QVBoxLayout, QWidget, QMessageBox, QProgressBar)
from PyQt5.QtCore import Qt

class PDFExtractorApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Extractor de Datos de Clientes")
        self.setGeometry(100, 100, 600, 400)
        self.output_file = "clientes.xlsx"
        self.initUI()

    def initUI(self):
        # Layout principal
        main_widget = QWidget()
        self.setCentralWidget(main_widget)
        layout = QVBoxLayout()
        main_widget.setLayout(layout)
        
        # Título
        title_label = QLabel("Extractor de Datos de Clientes")
        title_label.setAlignment(Qt.AlignCenter)
        title_label.setStyleSheet("font-size: 18px; font-weight: bold; margin: 10px;")
        layout.addWidget(title_label)
        
        # Descripción
        desc_label = QLabel("Esta aplicación extrae datos de clientes desde facturas PDF y los guarda en Excel.\n"
                           "Si el RUT ya existe en la base de datos, no se duplicará el registro.")
        desc_label.setAlignment(Qt.AlignCenter)
        desc_label.setWordWrap(True)
        layout.addWidget(desc_label)
        
        # Botón para seleccionar archivo PDF
        self.select_button = QPushButton("Seleccionar PDF")
        self.select_button.clicked.connect(self.select_pdf)
        self.select_button.setStyleSheet("font-size: 14px; padding: 10px; margin: 10px;")
        layout.addWidget(self.select_button)
        
        # Botón para seleccionar varios archivos PDF
        self.select_multiple_button = QPushButton("Seleccionar Múltiples PDFs")
        self.select_multiple_button.clicked.connect(self.select_multiple_pdfs)
        self.select_multiple_button.setStyleSheet("font-size: 14px; padding: 10px; margin: 10px;")
        layout.addWidget(self.select_multiple_button)
        
        # Etiqueta para mostrar archivo seleccionado
        self.file_label = QLabel("Ningún archivo seleccionado")
        self.file_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(self.file_label)
        
        # Barra de progreso
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        layout.addWidget(self.progress_bar)
        
        # Etiqueta para mostrar el estado
        self.status_label = QLabel("")
        self.status_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(self.status_label)
        
        # Botón para ver el archivo Excel
        self.view_excel_button = QPushButton("Ver Archivo Excel")
        self.view_excel_button.clicked.connect(self.open_excel)
        self.view_excel_button.setStyleSheet("font-size: 14px; padding: 10px; margin: 10px;")
        layout.addWidget(self.view_excel_button)

        # Información del creador
        creator_label = QLabel("Creador: Camilo Zavala - C1ZC<br>"
                                "Portafolio: <a href='https://c1zc.github.io/CamiloZavala/'>https://c1zc.github.io/CamiloZavala/</a>")
        creator_label.setAlignment(Qt.AlignCenter)
        creator_label.setOpenExternalLinks(True)  # Permitir abrir el enlace en el navegador
        creator_label.setStyleSheet("font-size: 12px; margin: 10px; color: gray;")
        layout.addWidget(creator_label)
        
        # Espaciador
        layout.addStretch()

    def select_pdf(self):
        """Abre un diálogo para seleccionar un archivo PDF"""
        file_path, _ = QFileDialog.getOpenFileName(self, "Seleccionar PDF", "", "PDF Files (*.pdf)")
        if file_path:
            self.file_label.setText(f"Archivo seleccionado: {os.path.basename(file_path)}")
            self.process_pdf(file_path)

    def select_multiple_pdfs(self):
        """Abre un diálogo para seleccionar múltiples archivos PDF"""
        file_paths, _ = QFileDialog.getOpenFileNames(self, "Seleccionar PDFs", "", "PDF Files (*.pdf)")
        if file_paths:
            self.file_label.setText(f"Archivos seleccionados: {len(file_paths)}")
            self.process_multiple_pdfs(file_paths)

    def process_multiple_pdfs(self, file_paths):
        """Procesa múltiples archivos PDF"""
        if not file_paths:
            return
            
        self.progress_bar.setVisible(True)
        self.progress_bar.setRange(0, len(file_paths))
        self.progress_bar.setValue(0)
        
        success_count = 0
        error_count = 0
        duplicate_count = 0
        
        for i, pdf_path in enumerate(file_paths):
            try:
                self.status_label.setText(f"Procesando {os.path.basename(pdf_path)}...")
                QApplication.processEvents()  # Actualiza la interfaz
                
                result = self.extract_and_save_data(pdf_path)
                if result == "success":
                    success_count += 1
                elif result == "duplicate":
                    duplicate_count += 1
                else:
                    error_count += 1
                    
                self.progress_bar.setValue(i + 1)
                QApplication.processEvents()
                
            except Exception as e:
                print(f"Error al procesar {pdf_path}: {str(e)}")
                error_count += 1
                self.progress_bar.setValue(i + 1)
                QApplication.processEvents()
        
        self.status_label.setText(f"Proceso completado: {success_count} archivos procesados, "
                                 f"{duplicate_count} duplicados, "
                                 f"{error_count} con errores")
        self.progress_bar.setVisible(False)

    def process_pdf(self, pdf_path):
        """Procesa un solo archivo PDF"""
        if not pdf_path:
            return
            
        self.status_label.setText("Procesando PDF...")
        QApplication.processEvents()  # Actualiza la interfaz
        
        try:
            result = self.extract_and_save_data(pdf_path)
            if result == "success":
                self.status_label.setText("¡Datos extraídos y guardados con éxito!")
            elif result == "duplicate":
                self.status_label.setText("Cliente ya existe en la base de datos. No se agregó.")
            else:
                self.status_label.setText("No se pudieron extraer todos los datos requeridos")
        except Exception as e:
            self.status_label.setText(f"Error: {str(e)}")
            QMessageBox.critical(self, "Error", f"Ocurrió un error al procesar el PDF: {str(e)}")

    def extract_and_save_data(self, pdf_path):
        """Extrae datos del PDF y los guarda en Excel si no existe el RUT"""
        # Extraer texto del PDF
        text = self.extract_text_from_pdf(pdf_path)
        if not text:
            return "error"
            
        # Extraer datos del texto
        data = self.extract_data_from_text(text)
        
        # Verificar si hay datos suficientes
        if not data["RUT"]:
            QMessageBox.warning(self, "Advertencia", "No se pudo encontrar el RUT del cliente en el documento.")
            return "error"
            
        # Validar si hay campos faltantes importantes
        missing_fields = [k for k, v in data.items() if not v]
        if missing_fields:
            QMessageBox.warning(self, "Advertencia", 
                              f"No se pudieron encontrar los siguientes campos: {', '.join(missing_fields)}")
        
        # Verificar si el cliente ya existe
        if self.client_exists(data["RUT"]):
            return "duplicate"
        
        # Guardar datos en Excel
        self.save_to_excel(data)
        return "success"

    def extract_text_from_pdf(self, pdf_path):
        """Extrae todo el texto de un PDF"""
        full_text = ""
        try:
            with pdfplumber.open(pdf_path) as pdf:
                for page in pdf.pages:
                    full_text += page.extract_text() + "\n"
            return full_text
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al leer el PDF: {str(e)}")
            return None

    def extract_data_from_text(self, text):
        """Extrae datos específicos del texto usando expresiones regulares"""
        data = {
            "Razón social": None,
            "RUT": None,
            "Giro": None,
            "Dirección": None,
            "Comuna": None,
            "Ciudad": None,
            "Nombre contacto": None,
            "Teléfono": None
        }
        
        # Mejorado: Primero identificamos claramente la sección del cliente
        # Buscamos el texto entre "SEÑOR(ES):" y algún otro indicador de fin de sección
        cliente_section_match = re.search(r'SEÑOR\(ES\):(.*?)(?=FACTURA|\bCOND\b|DETALLE|FECHA|FORMA DE PAGO|$)', 
                                         text, re.DOTALL | re.IGNORECASE)
        
        if cliente_section_match:
            cliente_text = cliente_section_match.group(1)
            
            # Extraer Razón Social
            razon_social_match = re.search(r'SEÑOR\(ES\):\s*(.+?)(?=\n|R\.U\.T\.|$)', text, re.IGNORECASE)
            if razon_social_match:
                data["Razón social"] = razon_social_match.group(1).strip()
            else:
                print("No se encontró la razón social en el texto extraído.")
                print("Texto extraído para depuración:")
                print(text)
            
            # Extraer RUT con formato específico
            rut_match = re.search(r'R\.U\.T\.:\s*([\d\.]+)-\s*([0-9kK])', cliente_text, re.IGNORECASE)
            if rut_match:
                # Capturar las partes del RUT
                cuerpo_rut = rut_match.group(1).replace('.', '')  # Eliminar puntos
                digito_verificador = rut_match.group(2)  # Capturar el dígito verificador (puede ser 'k' o un número)

                if digito_verificador:
                    # Formatear el RUT como xx.xxx.xxx-N
                    cuerpo_rut = f"{int(cuerpo_rut):,}".replace(",", ".")  # Agregar puntos como separadores de miles
                    data["RUT"] = f"{cuerpo_rut}-{digito_verificador.upper()}"  # Asegurar que el dígito verificador esté en mayúscula
                else:
                    print("Advertencia: No se encontró el dígito verificador del RUT.")
            else:
                print("No se encontró un RUT válido en el texto extraído.")
            
            # Extraer Giro
            giro_match = re.search(r'GIRO:\s*(.*?)(?=\n|DIRECC|$)', cliente_text, re.IGNORECASE)
            if giro_match:
                data["Giro"] = giro_match.group(1).strip()
            
            # Extraer Dirección
            direccion_match = re.search(r'DIRECCION:\s*(.*?)(?=\n|COMUNA|$)', cliente_text, re.IGNORECASE)
            if direccion_match:
                data["Dirección"] = direccion_match.group(1).strip()
            
            # Extraer Comuna
            comuna_match = re.search(r'COMUNA\s*(.*?)(?=\n|CIUDAD|$)', cliente_text, re.IGNORECASE)
            if comuna_match:
                data["Comuna"] = comuna_match.group(1).strip()
            
            # Extraer Ciudad
            ciudad_match = re.search(r'CIUDAD:\s*(.*?)(?=\n|CONTACTO|$)', cliente_text, re.IGNORECASE)
            if ciudad_match:
                data["Ciudad"] = ciudad_match.group(1).strip()
            
            # Extraer Nombre de contacto
            contacto_match = re.search(r'CONTACTO:\s*(.*?)(?=\n|F:|$)', cliente_text, re.IGNORECASE)
            if contacto_match:
                data["Nombre contacto"] = contacto_match.group(1).strip()
            
            # Extraer Teléfono y corregir formato
            telefono_match = re.search(
                r'F:\s*[:\-]?\s*(\d[\d\s\-]*)', cliente_text, re.IGNORECASE)
            if telefono_match:
                numero = telefono_match.group(1).strip()
                # Eliminar cualquier carácter no numérico excepto los dígitos
                numero = re.sub(r'\D', '', numero)
                if len(numero) == 9 and numero.startswith('9'):
                    # Formato correcto para números de 9 dígitos
                    data["Teléfono"] = f"+569 {numero[1:]}"
                elif len(numero) == 8:
                    # Formato correcto para números de 8 dígitos
                    data["Teléfono"] = f"+569 {numero}"
                else:
                    # Si el formato es diferente, guardar el número sin cambios
                    data["Teléfono"] = f"+56 {numero}"
        
        # Si no encontramos datos suficientes en la sección del cliente, hacer una búsqueda más general
        if not data["RUT"]:
            # Buscar secuencia SEÑOR(ES) -> RUT -> GIRO para identificar datos del cliente
            señores_pos = text.find("SEÑOR(ES):")
            if señores_pos != -1:
                text_after_señores = text[señores_pos:]
                
                # Extraer Razón Social si no se encontró antes
                if not data["Razón social"]:
                    razon_social_match = re.search(r'SEÑOR\(ES\):\s*(.*?)(?=\n|R\.U\.T\.)', text_after_señores, re.IGNORECASE)
                    if razon_social_match:
                        data["Razón social"] = razon_social_match.group(1).strip()
                
                # Extraer RUT si no se encontró antes
                if not data["RUT"]:
                    rut_match = re.search(r'R\.U\.T\.:\s*([\d\.\-]+)', text_after_señores, re.IGNORECASE)
                    if rut_match:
                        raw_rut = rut_match.group(1).strip()
                        clean_rut = re.sub(r'\s+', '', raw_rut)
                        clean_rut = re.sub(r'-\s+', '-', clean_rut)
                        data["RUT"] = clean_rut
                
                # Si encontramos RUT, buscar el resto de los datos a partir de ahí
                if data["RUT"]:
                    rut_pos = text_after_señores.find(data["RUT"])
                    if rut_pos != -1:
                        text_after_rut = text_after_señores[rut_pos:]
                        
                        # Extraer Giro si no se encontró antes
                        if not data["Giro"]:
                            giro_match = re.search(r'GIRO:\s*(.*?)(?=\n|DIRECC|$)', text_after_rut, re.IGNORECASE)
                            if giro_match:
                                data["Giro"] = giro_match.group(1).strip()
                        
                        # Extraer el resto de los campos si no se encontraron antes
                        if not data["Dirección"]:
                            direccion_match = re.search(r'DIRECCION:\s*(.*?)(?=\n|COMUNA|$)', text_after_rut, re.IGNORECASE)
                            if direccion_match:
                                data["Dirección"] = direccion_match.group(1).strip()
                        
                        if not data["Comuna"]:
                            comuna_match = re.search(r'COMUNA:\s*(.*?)(?=\n|CIUDAD|$)', text_after_rut, re.IGNORECASE)
                            if comuna_match:
                                data["Comuna"] = comuna_match.group(1).strip()
                        
                        if not data["Ciudad"]:
                            ciudad_match = re.search(r'CIUDAD:\s*(.*?)(?=\n|CONTACTO|$)', text_after_rut, re.IGNORECASE)
                            if ciudad_match:
                                data["Ciudad"] = ciudad_match.group(1).strip()
                        
                        if not data["Nombre contacto"]:
                            contacto_match = re.search(r'CONTACTO:\s*(.*?)(?=\n|F:|$)', text_after_rut, re.IGNORECASE)
                            if contacto_match:
                                data["Nombre contacto"] = contacto_match.group(1).strip()
                        
                        if not data["Teléfono"]:
                            telefono_match = re.search(r'F:\s*(\d+)', text_after_rut, re.IGNORECASE)
                            if telefono_match:
                                numero = telefono_match.group(1).strip()
                                # Formatear el número al formato internacional +569 XXXXXXXX
                                if len(numero) == 9 and numero.startswith('9'):
                                    data["Teléfono"] = f"+569 {numero[1:]}"
                                elif len(numero) == 8:
                                    data["Teléfono"] = f"+569 {numero}"
                                else:
                                    data["Teléfono"] = numero
        
        return data

    def client_exists(self, rut):
        """Verifica si el RUT ya existe en el archivo Excel"""
        if not os.path.exists(self.output_file):
            return False
            
        try:
            wb = load_workbook(self.output_file)
            ws = wb.active
            
            # Determinar el índice de la columna RUT
            rut_column = None
            for col_num, cell in enumerate(ws[1], 1):
                if cell.value == "RUT":
                    rut_column = col_num
                    break
                    
            if rut_column is None:
                return False
                
            # Buscar el RUT en la columna
            for row in range(2, ws.max_row + 1):
                cell_value = ws.cell(row=row, column=rut_column).value
                if cell_value == rut:
                    return True
                    
            return False
            
        except Exception as e:
            QMessageBox.warning(self, "Advertencia", 
                              f"Error al verificar duplicados: {str(e)}\n"
                              f"Continuando con el proceso...")
            return False

    def save_to_excel(self, data):
        """Guarda los datos en un archivo Excel"""
        # Verificar si el archivo existe
        if os.path.exists(self.output_file):
            try:
                wb = load_workbook(self.output_file)
                ws = wb.active
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Error al abrir el archivo Excel: {str(e)}")
                return
        else:
            # Crear nuevo archivo
            wb = Workbook()
            ws = wb.active
            # Añadir encabezados
            headers = list(data.keys())
            for col_num, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_num).value = header
        
        # Determinar la siguiente fila
        next_row = ws.max_row + 1
        if next_row == 2 and ws.cell(row=1, column=1).value is None:
            # Si el archivo está vacío, empezamos con la primera fila
            next_row = 1
            # Añadir encabezados
            headers = list(data.keys())
            for col_num, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_num).value = header
            next_row = 2
        
        # Añadir datos
        for col_num, (field, value) in enumerate(data.items(), 1):
            ws.cell(row=next_row, column=col_num).value = value
        
        # Guardar el archivo
        try:
            wb.save(self.output_file)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al guardar el archivo Excel: {str(e)}")

    def open_excel(self):
        """Abre el archivo Excel con la aplicación predeterminada"""
        if not os.path.exists(self.output_file):
            QMessageBox.information(self, "Información", "El archivo Excel aún no ha sido creado.")
            return
            
        # Abrir el archivo con la aplicación predeterminada
        try:
            if sys.platform == 'win32':
                os.startfile(self.output_file)
            elif sys.platform == 'darwin':  # macOS
                os.system(f'open "{self.output_file}"')
            else:  # Linux y otros
                os.system(f'xdg-open "{self.output_file}"')
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al abrir el archivo Excel: {str(e)}")

def main():
    app = QApplication(sys.argv)
    window = PDFExtractorApp()
    window.show()
    sys.exit(app.exec_())

if __name__ == "__main__":
    main()